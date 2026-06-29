"""Parse CatalogOne DG Excel workbooks (WLS Actions and Reasons format)."""

from __future__ import annotations

import io
import re
from dataclasses import dataclass, field
from typing import Any, BinaryIO

try:
    import openpyxl
except ImportError as exc:  # pragma: no cover - dependency guard
    raise ImportError("openpyxl is required for Excel DG import. pip install openpyxl") from exc

ACTION_ENTRY_SHEETS = frozenset(
    {
        "Add",
        "Cancel",
        "Change",
        "Terminate",
        "Suspend_Resume",
        "Re-establish",
        "Swap",
        "TOBR",
        "DRAFT_DEVICE",
    }
)

POLICY_SHEETS = frozenset(
    {
        "Default RC Calculation Strategy",
        "NoProrateAddon Strategy",
    }
)

REFERENCE_SHEETS = frozenset(
    {
        "Changes",
        "Business Owners",
        "Allow Action per State",
        "Backdate_future date",
        "Sheet1",
        "Sheet2",
        "Sheet3",
    }
)

_HEADER_ALIASES = {
    "reason_code": ("Reason_Code", "REASONCODE"),
    "reason_description": (
        "Reason_Description",
        "Reason_Code_Description",
        "DESCRIPTION",
    ),
    "order_action": ("Order_Action (Localized Name)", "Order_Action\n(Localized Name)"),
    "product_type": ("productType", "Product (internal use)"),
    "action_type": ("Action Type (Internal use)",),
    "request_id": ("Request ID", "Feature/US or Intake #", "Project ID (internal use)"),
}


@dataclass(frozen=True)
class DgRow:
    sheet: str
    category: str  # modify_reason | action | policy | reference
    reason_code: str
    reason_description: str
    order_action: str
    product_type: str
    request_id: str
    attributes: dict[str, Any] = field(default_factory=dict)


def _norm_header(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).replace("\n", " ")).strip()


def _find_header_row(rows: list[tuple[Any, ...]]) -> tuple[int, list[str]] | None:
    for index, row in enumerate(rows[:8]):
        headers = [_norm_header(cell) for cell in row]
        if any(
            header in headers or any(alias in headers for alias in group)
            for group in _HEADER_ALIASES.values()
            for header in (group if isinstance(group, tuple) else (group,))
        ):
            return index, headers
    return None


def _header_index(headers: list[str], *aliases: str) -> int | None:
    for alias in aliases:
        normalized = _norm_header(alias)
        for index, header in enumerate(headers):
            if header == normalized:
                return index
    return None


def _cell(row: tuple[Any, ...], index: int | None) -> Any:
    if index is None or index >= len(row):
        return None
    return row[index]


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return "" if text.upper() in {"", "NA", "N/A", "NONE"} else text


def _sheet_category(sheet_name: str) -> str:
    if sheet_name == "Modify_Reasons":
        return "modify_reason"
    if sheet_name in ACTION_ENTRY_SHEETS:
        return "action"
    if sheet_name in POLICY_SHEETS:
        return "policy"
    return "reference"


def _parse_policy_sheet(sheet_name: str, rows: list[tuple[Any, ...]]) -> list[DgRow]:
    parsed: list[DgRow] = []
    for row in rows[2:]:
        if not any(row):
            continue
        policy_name = _clean_text(row[0] if len(row) > 0 else "")
        action = _clean_text(row[3] if len(row) > 3 else "")
        reason = _clean_text(row[4] if len(row) > 4 else "")
        directive = _clean_text(row[5] if len(row) > 5 else "")
        if not policy_name and not (action and reason):
            continue
        code = reason or policy_name
        parsed.append(
            DgRow(
                sheet=sheet_name,
                category="policy",
                reason_code=code,
                reason_description=directive or policy_name,
                order_action=action,
                product_type=_clean_text(row[6] if len(row) > 6 else ""),
                request_id="",
                attributes={
                    "policyName": policy_name,
                    "localizedPolicyName": _clean_text(row[1] if len(row) > 1 else ""),
                    "prorationDirective": directive,
                },
            )
        )
    return parsed


def _parse_data_sheet(sheet_name: str, rows: list[tuple[Any, ...]]) -> list[DgRow]:
    category = _sheet_category(sheet_name)
    if category == "policy":
        return _parse_policy_sheet(sheet_name, rows)

    header_info = _find_header_row(rows)
    if not header_info:
        return []

    header_index, headers = header_info
    idx_code = _header_index(headers, *_HEADER_ALIASES["reason_code"])
    idx_desc = _header_index(headers, *_HEADER_ALIASES["reason_description"])
    idx_action = _header_index(headers, *_HEADER_ALIASES["order_action"])
    idx_product = _header_index(headers, *_HEADER_ALIASES["product_type"])
    idx_action_type = _header_index(headers, *_HEADER_ALIASES["action_type"])
    idx_request = _header_index(headers, *_HEADER_ALIASES["request_id"])

    parsed: list[DgRow] = []
    for row in rows[header_index + 1 :]:
        if not any(row):
            continue
        reason_code = _clean_text(_cell(row, idx_code))
        if not reason_code:
            continue

        attributes: dict[str, Any] = {}
        for col_index, header in enumerate(headers):
            if not header or col_index == idx_code:
                continue
            value = _cell(row, col_index)
            if value is None or str(value).strip() == "":
                continue
            attributes[header] = value

        order_action = _clean_text(_cell(row, idx_action))
        if not order_action and idx_action_type is not None:
            order_action = _clean_text(_cell(row, idx_action_type))
        if not order_action and category == "action":
            order_action = sheet_name.lower().replace("_", " ")

        parsed.append(
            DgRow(
                sheet=sheet_name,
                category=category,
                reason_code=reason_code,
                reason_description=_clean_text(_cell(row, idx_desc)),
                order_action=order_action,
                product_type=_clean_text(_cell(row, idx_product)),
                request_id=_clean_text(_cell(row, idx_request)),
                attributes=attributes,
            )
        )
    return parsed


def parse_excel_workbook(source: bytes | BinaryIO) -> tuple[list[DgRow], list[str]]:
    """Return parsed DG rows and workbook sheet names."""
    if isinstance(source, bytes):
        source = io.BytesIO(source)

    workbook = openpyxl.load_workbook(source, read_only=True, data_only=True)
    sheet_names = list(workbook.sheetnames)
    rows_by_sheet: list[DgRow] = []

    for sheet_name in sheet_names:
        worksheet = workbook[sheet_name]
        rows = [tuple(row) for row in worksheet.iter_rows(values_only=True)]
        rows_by_sheet.extend(_parse_data_sheet(sheet_name, rows))

    workbook.close()
    return rows_by_sheet, sheet_names
